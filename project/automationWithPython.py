import openpyxl as xl
wb=xl.load_workbook(r'D:\python from scratch\project\transactions.xlsx')
sheet = wb["Sheet1"]
cell=sheet['a1']
cell = sheet.cell(1,1)
for row in range(1, sheet.max_row + 1):
    print(row)